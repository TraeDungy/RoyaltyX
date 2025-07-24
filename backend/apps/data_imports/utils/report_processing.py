import csv
import hashlib
import io
import logging
import re
from datetime import datetime
from decimal import Decimal
from typing import Any, BinaryIO, Dict, List, Optional, Tuple

import openpyxl

from apps.product.models import (
    Product,
    ProductImpressions,
    ProductSale,
    ProductMetric,
)

logger = logging.getLogger(__name__)

COLUMN_ALIASES = {
    "Title": ["Title", "title", "program_name", "Title Name"],
    "Unit Price": ["Unit Price", "unit_price", "price"],
    "Unit Price Currency": ["Unit Price Currency", "unit_price_currency"],
    "Quantity": ["Quantity", "qty"],
    "Consumption Type": ["Consumption Type", "consumption_type"],
    "Is Refund": ["Is Refund", "is_refund"],
    "Royalty Amount": ["Royalty Amount", "royalty_amount"],
    "Royalty Currency": ["Royalty Currency", "royalty_currency"],
    "Period Start": ["Period Start", "period_start"],
    "Period End": ["Period End", "period_end"],
    "impressions": ["impressions"],
    "ecpm": ["ecpm", "eCPM"],
}

# Fields we already handle when importing reports
KNOWN_FIELDS = {
    "Title",
    "Unit Price",
    "Unit Price Currency",
    "Quantity",
    "Consumption Type",
    "Is Refund",
    "Royalty Amount",
    "Royalty Currency",
    "Period Start",
    "Period End",
    "impressions",
    "ecpm",
}


MONTH_NAMES = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "sept": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def parse_date_string(text: str) -> Optional[datetime]:
    """Attempt to parse a date from a string."""
    text = str(text).strip()
    try:
        return datetime.fromisoformat(text)
    except Exception:
        pass

    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y-%m",
        "%Y_%m",
        "%b %Y",
        "%B %Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def parse_month_year_from_filename(filename: str) -> Optional[Tuple[int, int]]:
    """Extract month and year information from a filename."""
    name = filename.lower()

    match = re.search(r"(20\d{2})[\-_\. ]?(\d{1,2})", name)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        if 1 <= month <= 12:
            return month, year

    match = re.search(r"(\d{1,2})[\-_\. ]?(20\d{2})", name)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if 1 <= month <= 12:
            return month, year

    match = re.search(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*[\-_ ]*(20\d{2})",
        name,
    )
    if match:
        month = MONTH_NAMES[match.group(1)[:3]]
        year = int(match.group(2))
        return month, year

    match = re.search(
        r"(20\d{2})[\-_ ]*(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*",
        name,
    )
    if match:
        year = int(match.group(1))
        month = MONTH_NAMES[match.group(2)[:3]]
        return month, year

    return None


def detect_delimiter(file: BinaryIO) -> str:
    """Try to detect the delimiter used in the CSV file."""
    file.seek(0)
    sample = file.read(1024).decode("utf-8")
    file.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t"])
        return dialect.delimiter
    except Exception:
        return ","


def map_header(header: str) -> str:
    for canonical, aliases in COLUMN_ALIASES.items():
        if header in aliases:
            return canonical
    return header


def detect_headers(file: BinaryIO) -> List[str]:
    """Return normalized header names from the first row of the file."""
    filename = getattr(file, "name", "").lower()
    headers: List[str] = []
    if filename.endswith((".xlsx", ".xls")):
        workbook = openpyxl.load_workbook(file, data_only=True)
        file.seek(0)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            headers = [map_header(str(h)) for h in rows[0]]
    else:
        delimiter = detect_delimiter(file)
        decoded_file = io.StringIO(file.read().decode("utf-8"))
        file.seek(0)
        reader = csv.reader(decoded_file, delimiter=delimiter)
        headers = next(reader, [])
        headers = [map_header(h) for h in headers]
    file.seek(0)
    return headers


def header_signature(headers: List[str]) -> str:
    """Create a deterministic signature for a set of headers."""
    normalized = [str(h).strip().lower() for h in headers if h is not None]
    return hashlib.sha256(",".join(sorted(normalized)).encode("utf-8")).hexdigest()



def validate_csv(file: BinaryIO) -> bool:
    """Validates if the uploaded file is a valid CSV."""
    try:
        delimiter = detect_delimiter(file)
        decoded_file = io.StringIO(file.read().decode("utf-8"))
        file.seek(0)
        reader = csv.reader(decoded_file, delimiter=delimiter)
        headers = next(reader, None)
        if not headers:
            return False
        return True
    except Exception as e:
        logger.error("CSV validation error: %s", e)
        return False


def read_csv(
    file: BinaryIO, custom_mapping: Dict[str, str] | None = None
) -> List[Dict[str, str]]:
    """Reads a CSV file and returns its content as a list of dictionaries.

    Parameters
    ----------
    file: BinaryIO
        The file object to read from.
    custom_mapping: Dict[str, str] | None
        Optional mapping from CSV headers to canonical column names.
    """
    delimiter = detect_delimiter(file)
    file.seek(0)  # Ensure the file pointer is at the beginning
    decoded_file = io.StringIO(file.read().decode("utf-8"))
    reader = csv.DictReader(decoded_file, delimiter=delimiter)

    headers_map = {h: map_header(h) for h in reader.fieldnames}
    if custom_mapping:
        headers_map.update(custom_mapping)

    data = []
    for row in reader:
        normalized = {headers_map.get(k, k): v for k, v in row.items()}
        data.append(normalized)
    return data


def validate_excel(file: BinaryIO) -> bool:
    """Validate that the uploaded file is a valid Excel workbook."""
    try:
        openpyxl.load_workbook(file)
        file.seek(0)
        return True
    except Exception as e:  # pragma: no cover - log the error for debugging
        print(f"Excel validation error: {e}")
        return False


def read_excel(
    file: BinaryIO, custom_mapping: Dict[str, str] | None = None
) -> List[Dict[str, str]]:
    """Read an Excel file and return its content as a list of dictionaries."""
    workbook = openpyxl.load_workbook(file, data_only=True)
    file.seek(0)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [map_header(str(h)) for h in rows[0]]
    if custom_mapping:
        headers = [custom_mapping.get(h, h) for h in headers]

    data: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        normalized = {headers[i]: row[i] for i in range(len(headers))}
        data.append(normalized)
    return data


def read_report(
    file: BinaryIO, custom_mapping: Dict[str, str] | None = None
) -> List[Dict[str, str]]:
    """Read a report file in CSV or Excel format."""
    filename = getattr(file, "name", "").lower()
    if filename.endswith(('.xlsx', '.xls')):
        return read_excel(file, custom_mapping=custom_mapping)
    return read_csv(file, custom_mapping=custom_mapping)


def update_products(
    data: List[Dict[str, str]], project_id: int, file_id: int
) -> Dict[str, int]:
    """Updates products based on CSV data and returns a summary."""
    updated_count = 0

    for row in data:
        title = row.get("Title")
        if not title:
            continue

        product = Product.objects.filter(title=title, project_id=project_id).first()
        if not product:
            product = Product.objects.create(title=title, project_id=project_id)

        if row.get("Unit Price"):
            storeProductSales(row, product, file_id)

        if row.get("impressions"):
            storeProductImpressions(row, product, file_id)

        storeProductMetrics(row, product, file_id, KNOWN_FIELDS)

        updated_count += 1

    return {"updated": updated_count}


def process_report(
    file: BinaryIO,
    project_id: int,
    file_id: int,
    column_mapping: Dict[str, str] | None = None,
) -> Dict[str, str]:
    """Process a report file (CSV or Excel) and update products."""
    try:
        filename = getattr(file, "name", "").lower()
        if filename.endswith((".xlsx", ".xls")):
            if not validate_excel(file):
                return {"status": "error", "message": "Invalid Excel file"}
        else:
            if not validate_csv(file):
                return {"status": "error", "message": "Invalid CSV file"}

        data = read_report(file, custom_mapping=column_mapping)
        result = update_products(data, project_id, file_id)

        return {
            "status": "success",
            "message": f"Updated {result['updated']} products",
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


def storeProductSales(row: Dict[str, Any], product: Product, file_id: int) -> None:
    ProductSale.objects.create(
        product=product,
        type=(row.get("Consumption Type") or "").lower(),
        unit_price=Decimal(row.get("Unit Price")),
        unit_price_currency=row.get("Unit Price Currency"),
        quantity=Decimal(row.get("Quantity")),
        is_refund=row.get("Is Refund") == "Yes",
        royalty_amount=Decimal(row.get("Royalty Amount")),
        royalty_currency=row.get("Royalty Currency"),
        period_start=row.get("Period Start"),
        period_end=row.get("Period End"),
        from_file_id=file_id,
    )


def storeProductImpressions(
    row: Dict[str, Any], product: Product, file_id: int
) -> None:
    try:
        ProductImpressions.objects.create(
            product=product,
            ecpm=Decimal(row.get("ecpm")) if row.get("ecpm") else None,
            impressions=row.get("impressions"),
            period_start=row.get("Period Start"),
            period_end=row.get("Period End"),
            from_file_id=file_id,
        )
    except Exception as e:
        logger.error("Failed to store product impressions: %s", e)


def storeProductMetrics(
    row: Dict[str, Any], product: Product, file_id: int, known_fields: set
) -> None:
    for key, value in row.items():
        if key in known_fields:
            continue
        try:
            num = Decimal(str(value))
        except Exception:
            continue
        try:
            ProductMetric.objects.create(
                product=product,
                name=key,
                value=num,
                period_start=row.get("Period Start"),
                period_end=row.get("Period End"),
                from_file_id=file_id,
            )
        except Exception as e:
            logger.error("Failed to store metric %s: %s", key, e)
