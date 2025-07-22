import csv
import io
from typing import BinaryIO, Dict, List

from openpyxl import load_workbook

from apps.product.models import Product, ProductUser
from apps.user.models import User


def detect_file_type(file: BinaryIO) -> str:
    """Return 'xlsx' if the uploaded file is an Excel file, otherwise 'csv'."""
    name = getattr(file, "name", "").lower()
    if name.endswith(".xlsx"):
        return "xlsx"
    return "csv"


def validate_file(file: BinaryIO) -> bool:
    """Validates if the uploaded file contains headers."""
    file_type = detect_file_type(file)
    try:
        if file_type == "xlsx":
            workbook = load_workbook(file, read_only=True)
            file.seek(0)
            sheet = workbook.active
            headers = next(sheet.iter_rows(values_only=True), None)
            return bool(headers)
        else:
            decoded_file = io.StringIO(file.read().decode("utf-8"))
            file.seek(0)
            reader = csv.reader(decoded_file)
            headers = next(reader, None)
            return bool(headers)
    except Exception as e:
        print(f"File validation error: {e}")
        return False


def read_csv(file: BinaryIO) -> List[Dict[str, str]]:
    """Reads a CSV file and returns its content as a list of dictionaries."""
    file.seek(0)
    decoded_file = io.StringIO(file.read().decode("utf-8"))
    reader = csv.DictReader(decoded_file)
    return [row for row in reader]


def read_xlsx(file: BinaryIO) -> List[Dict[str, str]]:
    """Reads an XLSX file and returns its content as a list of dictionaries."""
    file.seek(0)
    workbook = load_workbook(file, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data: List[Dict[str, str]] = []
    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_dict)
    return data


def read_file(file: BinaryIO) -> List[Dict[str, str]]:
    """Read uploaded file regardless of type."""
    return read_xlsx(file) if detect_file_type(file) == "xlsx" else read_csv(file)


def update_product_users(data: List[Dict[str, str]], project_id: int):
    """Updates products based on CSV data and returns a summary."""

    for row in data:
        title = row.get("title")
        email = row.get("email")
        producer_fee = row.get("producer fee")

        user = User.objects.filter(email=email).first()

        if not title or not email or not producer_fee or not user:
            continue

        product = Product.objects.filter(title=title, project_id=project_id).first()
        if not product:
            product = Product.objects.create(title=title, project_id=project_id)

        productUser = ProductUser.objects.filter(user=user, product=product).first()
        if not productUser:
            ProductUser.objects.create(
                product=product, user=user, producer_fee=int(producer_fee)
            )

    return {"message": "success"}


def process_producers(file: BinaryIO, project_id: int) -> Dict[str, str]:
    """Processes file with list of producers and updates products."""
    try:
        if not validate_file(file):
            return {"status": "error", "message": "Invalid file"}

        data = read_file(file)
        result = update_product_users(data, project_id)

        return {
            "status": "success",
            "message": result,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
